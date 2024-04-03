#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 14 09:40:55 2024

@author: Rui
"""
import os
import re
import gzip
import goatools
import numpy as np
import pandas as pd
import urllib.request
from collections import defaultdict
from itertools import combinations
import seaborn as sns
import matplotlib.pyplot as plt
import omicverse as ov
import gseapy as gp
import scipy.cluster.hierarchy as shc
from sklearn.decomposition import PCA
from matplotlib_venn import venn2
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#——————————————数据处理——————————————#
def matrix_open_and_dispose(aaa):
    with gzip.open(aaa,"r") as file:
        content = file.readlines()
    for i in content:
        if "!series_matrix_table_begin" in str(i):
            temp_location = content.index(i)
        if "!Sample_title" in str(i):
            title_str = str(i)[2:-3]
        if "!Series_geo_accession" in str(i):
            geo_id = str(re.findall(r'"([^"]+)"', str(i)[2:-3]))[2:-2]
        if '!Sample_characteristics_ch1' in str(i):
            group_info = str(i).split('\\')
    # 原始字符串
    original_string = title_str
    title_list = re.findall(r'"([^"]+)"', original_string) 
    
    temp_location +=1
    with gzip.open(aaa,"r") as file:
        gene_expression_dataframe = pd.read_csv(file, delimiter='\t', skiprows=temp_location, names=title_list)
    
    gene_expression_dataframe = gene_expression_dataframe.drop(gene_expression_dataframe.index[-1])
    gene_expression_dataframe = gene_expression_dataframe.drop(gene_expression_dataframe.index[0])
    gene_expression_dataframe = gene_expression_dataframe.reset_index(drop=False)
    gene_expression_dataframe = gene_expression_dataframe.rename(columns={"index":"ID"})
    gene_expression_dataframe["ID"] = gene_expression_dataframe["ID"].astype(str)
    for col in gene_expression_dataframe.columns[1:]:
        gene_expression_dataframe[col] = gene_expression_dataframe[col].astype(float)
    return gene_expression_dataframe, group_info


def download_GPL(platform_id, save_path_GPL):
    try:
        temp_url = f'ftp://ftp.ncbi.nlm.nih.gov/geo/platforms/{platform_id[:-3]}nnn/{platform_id}/annot/{platform_id}.annot.gz'
        save_path_GPL += f"{platform_id}.annot.gz"
        # 打开FTP链接
        with urllib.request.urlopen(temp_url) as response:
            # 读取数据
            data = response.read()
            # 将数据写入到本地文件
            with open(save_path_GPL, 'wb') as f:
                f.write(data)
        #print("基因ID注释文件下载成功")
    except Exception as e:
        pass
        #print("基因ID注释文件下载失败:", e)
# 调用函数下载文件
def open_annot_file(platform_id, save_path_GPL):
    try:
        path = save_path_GPL + f"{platform_id}.annot.gz"
        #print(path)
        #print(save_path_GPL)
        with gzip.open(path, 'rt') as f:  # 'rt'表示以文本模式读取压缩文件
            # 读取文件内容
            content = f.readlines()
            header_index = 0
            for i, line in enumerate(content):
                if '\t' in line:
                    header_index = i
                    break
            ID_conversion_df = pd.read_csv(path, sep='\t', skiprows=header_index)
            ID_conversion_df = ID_conversion_df.drop(ID_conversion_df.index[-1])
            ID_conversion_df = ID_conversion_df.set_index(ID_conversion_df.columns[0])
            ID_conversion_df = ID_conversion_df.reset_index(drop=False)
        return ID_conversion_df
    except Exception as e:
        try:
            path = save_path_GPL + f"{platform_id}.annot.gz"
            with open(path, 'rt') as f: 
                content = f.readlines()
                header_index = 0
                for i, line in enumerate(content):
                    if '\t' in line:
                        header_index = i
                        break
                ID_conversion_df = pd.read_csv(path, sep='\t', skiprows=header_index)
                ID_conversion_df = ID_conversion_df.drop(ID_conversion_df.index[-1])
                ID_conversion_df = ID_conversion_df.set_index(ID_conversion_df.columns[0])
                ID_conversion_df = ID_conversion_df.reset_index(drop=False)
            return ID_conversion_df
        except:
            print("打开文件失败:", e)
            return None


#——————————————GO分析的前置操作——————————————#
def GO_pre(obo_fname_path, fin_gene2go_path, Taxonomy_list):
    from goatools.base import download_go_basic_obo
    from goatools.base import download_ncbi_associations
    from goatools.obo_parser import GODag
    from goatools.anno.genetogo_reader import Gene2GoReader
    obo_fname = download_go_basic_obo()
    fin_gene2go = download_ncbi_associations()
    obodag = GODag(obo_fname)
    objanno = Gene2GoReader(fin_gene2go, taxids=Taxonomy_list)#
    ns2assoc = objanno.get_ns2assc()
    for nspc, id2gos in ns2assoc.items():
        print("{NS} {N:,} annotated human genes".format(NS=nspc, N=len(id2gos)))
    
    import sys
    sys.path.append('/Users/Rui/anaconda3/envs/Bio_analysis')
    
    from genes_ncbi_human_proteincoding import GENEID2NT as GeneID2nt_human
    #print(len(GeneID2nt_human))#检测是否正常运行
    from goatools.goea.go_enrichment_ns import GOEnrichmentStudyNS
    goeaobj = GOEnrichmentStudyNS(GeneID2nt_human.keys(), 
                                  ns2assoc, 
                                  obodag, 
                                  propagate_counts=False, 
                                  alpha=0.05, 
                                  methods=['fdr_bh'])
    GO_items=[]#获取GEO文库中的BP、CC、MF基因编号
    """Cellular component，CC 细胞成分; Biological process，BP 生物学过程; Molecular function，MF 分子功能"""
    temp = goeaobj.ns2objgoea["BP"].assoc
    for i in temp:
        GO_items += temp[i]
    temp = goeaobj.ns2objgoea["CC"].assoc
    for i in temp:
        GO_items += temp[i]
    temp = goeaobj.ns2objgoea["MF"].assoc
    for i in temp:
        GO_items += temp[i]
    
    id_mapper = {}
    for key in GeneID2nt_human:
        id_mapper[GeneID2nt_human[key].Symbol] = GeneID2nt_human[key].GeneID
    rev_id_mapper = {v:k for k, v in id_mapper.items()}
    return id_mapper, rev_id_mapper, GO_items, goeaobj

#----封装函数----
def GO_analysis(gene_list, id_mapper, goeaobj, GO_items, rev_id_mapper):
    mapped_gene=[]
    for i in gene_list:
        try:
            mapped_gene.append(id_mapper[i])
        except:
            pass
    goea_result_all=goeaobj.run_study(mapped_gene)
    goea_result_sig=[x for x in goea_result_all if x.p_fdr_bh<0.01]
    df_go=pd.DataFrame(list(map(lambda x: [x.GO, x.goterm.name, x.goterm.namespace, x.p_uncorrected, x.p_fdr_bh,\
                         x.ratio_in_study[0], x.ratio_in_study[1], GO_items.count(x.GO), list(map(lambda y: rev_id_mapper[y], 
                         x.study_items))], goea_result_sig)), 
                         columns = ['GO', 'term', 'class', 'p', 'p_corr', 'n_genes','n_study', 'n_go', 'study_genes'])
    df_go = df_go[df_go.n_genes > 1]
    df_go['LogP'] = -np.log10(df_go['p'])
    return df_go
    
    
#——————————————info——————————————#
def Basic_pretreat(xiangxi, matrix_path, message_check_df, save_path_GPL, analysis_save):
    # 调用函数打开文件
    matrix_table, group_info = matrix_open_and_dispose(matrix_path)
    message_check_df_1 = message_check_df[message_check_df["GSE"] == xiangxi]
    GPL_NAME = message_check_df_1['GPL'].iloc[0].strip('[]').strip('\'')
    
    group_info = group_info[1:-1]
    group_info = [item[2:-1] for item in group_info]
    

    if not os.path.exists(save_path_GPL+f"{GPL_NAME}.annot.gz"):
        download_GPL(GPL_NAME, save_path_GPL)
    
    while True:
        ID_conversion_df = open_annot_file(GPL_NAME, save_path_GPL)
        
        if ID_conversion_df is None:
            print("您的网络连接或其他服务可能出现问题，请由GEO平台上手动下载注释文件。\n请确保它能够正确的解压，之后将其移动到您先前给出的目录")
            temp = save_path_GPL
            print(f"路径是:{temp}")
            print(f"所需注释文件的平台是:{xiangxi}")
            flag = input("我已将文件移动至对应位置(y/n/give_up)")
            if flag=="y":
                ID_conversion_df = open_annot_file(GPL_NAME, save_path_GPL)
            if flag == 'give_up':
                break
            else:
                print("\n请确保您已经将文件移动至对应位置")
        
        
        
        else:
            break
    # 定义正则表达式模式，用于匹配包含genesymbol的字符串，不区分大小写
    pattern1 = re.compile(r'genesymbol', re.IGNORECASE)
    pattern2 = re.compile(r'geneaccession', re.IGNORECASE)
    # 遍历DataFrame的列标签
    for column in ID_conversion_df.columns:
        # 将列标签中的下划线替换为空格，并转换为小写
        column = column.replace('_', ' ').lower()
        # 检查列标签是否包含 "genesymbol" 的内容
        if pattern1.search(column):
            # 假设 df 是你的 DataFrame，old_name 是原列名，new_name 是你想要改成的新列名
            ID_conversion_df.rename(columns={column: 'Gene symbol'}, inplace=True)
        if pattern2.search(column):
            ID_conversion_df['Gene symbol'] = ID_conversion_df[column].str.split('///').str[1]
 
    df_merged = pd.merge(matrix_table, ID_conversion_df[['ID','Gene symbol']], how="left")#将平台信息与矩阵信息依照ID_REF信息合并
    df_merged["Value_mean"]=matrix_table.iloc[:,1:].apply(lambda x:x.mean(),axis=1)
    df_temp = df_merged.groupby("Gene symbol",as_index=False).apply(lambda X:X.nlargest(1,"Value_mean"))#分组，并且每组保留最大值
    
    #生成处理完毕的表达矩阵
    expset = df_temp.drop(columns=["Value_mean"])
    expset.reset_index()
    expset.set_index("Gene symbol",inplace=True)
    
    # 删除第一列和最后一列
    new_expset = expset
    new_expset = new_expset.iloc[:, 1:]
    # 将 NaN 替换为 0
    new_expset = new_expset.fillna(0)

    iiii = new_expset.loc['GAPDH'][0]
    if iiii > 1000:
        new_expset = np.log2(new_expset) # 给 DataFrame 中每一项取 log2
        
        
    my_dict = defaultdict(list)
    keys = group_info
    values = list(new_expset.columns)
    for key, value in zip(keys, values):
        my_dict[key].append(value)
    my_dict = dict(my_dict)
    
    return my_dict, new_expset
    #重大bug，不止一种分组时，怎么办
def DEG_analysis(xiangxi, analysis_save, new_expset, my_dict):
    try:
        os.makedirs(analysis_save + xiangxi + '/' + "DEG/")
    except:
        pass
    #DEG分析
    FoldChange_threshold = 0.05
    qvalue_threshold = 0.05
    plt.figure(figsize=(8, 6))
    # 定义五个不同的组别
    stages = my_dict.keys()
    # 生成所有可能的两两组合
    max_q_mean = float('-inf')#用于筛选差异最大的组的预设q
    stage_combinations = list(combinations(stages, 2))
    for combination in stage_combinations:
        control_groups   = my_dict[combination[0]]
        treatment_groups = my_dict[combination[1]]
        temp = control_groups + treatment_groups
        new_expset_combination = new_expset[temp]
        
        ov.utils.ov_plot_set()
        dds=ov.bulk.pyDEG(new_expset_combination)#用自己的矩阵，新建一个差异表达分析的项目
        dds.drop_duplicates_index()
        #print('... drop_duplicates_index success')
        dds.normalize()#标准化处理数据
        #print('... estimateSizeFactors and normalize success')
    
        result=dds.deg_analysis(treatment_groups,control_groups,method='ttest')
        #result.head()
        #print(result.shape)
        result=result.loc[result['log2(BaseMean)']>1]
        #print(result.shape)
        # -1 means automatically calculates
        dds.foldchange_set(fc_threshold=-1,
                           pval_threshold=0.05,
                           logp_max=6)
        # NFE2L2 'MFN1','MFN2','FIS1','DRP1'
        #top_10           = new_expset_combination.loc[result['FoldChange'].nlargest(25).index]
        nfe2_start_rows  = result.filter(like='NFE2', axis=0)    
        mfn1_start_rows  = result.filter(like='MFN1', axis=0)
        mfn2_start_rows  = result.filter(like='MFN2', axis=0)
        fis_start_rows   = result.filter(like='FIS', axis=0)
        drp_start_rows   = result.filter(like='DRP', axis=0)
        
        best_find_df = pd.concat([nfe2_start_rows, mfn1_start_rows, mfn2_start_rows, fis_start_rows, drp_start_rows])   
        q_mean = best_find_df['qvalue'].mean()
        #q_median = median(list(Heatmap_df['qvalue']))
        #如果想改用中位数取最大差异组
        #from statistics import median
        if q_mean > max_q_mean:
            max_q_mean = q_mean
            best_combination = combination
            #best_combination_VP = plt
            best_result = result
        
    # 火山图绘制
    # new_expset_combination 是你的数据框，包含 log2FC 和 -log10(qvalue) 两列数据
    df_v=best_result
    df_v['color'] = 'grey'  # 默认灰色
    df_v.loc[(df_v['FoldChange'] > 1+FoldChange_threshold)&(df_v['-log(pvalue)'] > qvalue_threshold), 'color'] = 'red'  # 红色点
    df_v.loc[(df_v['FoldChange'] < 1-FoldChange_threshold)&(df_v['-log(pvalue)'] > qvalue_threshold), 'color'] = 'blue'  # 蓝色点
    #df.loc[(df['-log(qvalue)'] < y_threshold)&(df['log2FC'] < 0.5)&(df['log2FC'] > -0.5), 'color'] = 'grey'  # 灰色点
    sns.scatterplot(data=df_v.loc[(df_v['color']=='red')], 
                    x='FoldChange', y='-log(pvalue)', 
                    hue='color',palette=['red'])
    sns.scatterplot(data=df_v.loc[(df_v['color']=='blue')], 
                    x='FoldChange', y='-log(pvalue)', 
                    hue='color',palette=['blue'])
    sns.scatterplot(data=df_v.loc[(df_v['color']=='grey')], 
                    x='FoldChange', y='-log(pvalue)', 
                    hue='color',palette=['grey'])
    # 添加阈值线
    plt.axvline(x=1+FoldChange_threshold, color='black', linestyle='--', linewidth=1)
    plt.axvline(x=1-FoldChange_threshold, color='black', linestyle='--', linewidth=1)
    plt.axhline(y=qvalue_threshold, color='black', linestyle='--', linewidth=1)
    # 添加标题和标签
    plt.title('Volcano Plot')
    plt.xlabel('FoldChange')
    plt.ylabel('-log10(p-value)')
    # 显示图形
    red_patch = plt.Line2D([0], [0], marker='o', color='w', label='Upregulated', markersize=10, markerfacecolor='red') 
    blue_patch = plt.Line2D([0], [0], marker='o', color='w', label='Downregulated', markersize=10, markerfacecolor='blue') 
    plt.legend(handles=[red_patch, blue_patch])
    plt.title('Volcano Plot of Different Expression Genes')
    plt.savefig(analysis_save + xiangxi + '/' + "DEG/" + str(best_combination)[2:-2] +'_Volcano_plot.tiff',dpi=300,bbox_inches='tight')
    plt.show()
    
    control_groups   = my_dict[best_combination[0]]
    treatment_groups = my_dict[best_combination[1]]
    new_expset_combination_1 = new_expset[control_groups]
    new_expset_combination_2 = new_expset[treatment_groups]
    new_expset_combination = pd.concat([new_expset_combination_1, new_expset_combination_2], axis=1)
    
    #热图绘制
    #top_10 = best_result['qvalue'].nlargest(10).index
    new_df = new_expset_combination[new_expset_combination.index.str.len() <= 10].head(10)
    nfe2_start_rows  = new_expset_combination.filter(like='NFE2', axis=0)    
    mfn1_start_rows  = new_expset_combination.filter(like='MFN1', axis=0)
    mfn2_start_rows  = new_expset_combination.filter(like='MFN2', axis=0)
    fis_start_rows   = new_expset_combination.filter(like='FIS', axis=0)
    drp_start_rows   = new_expset_combination.filter(like='DRP', axis=0)
    new_df = pd.concat([new_df, nfe2_start_rows, mfn1_start_rows, mfn2_start_rows, fis_start_rows, drp_start_rows])
    
    plt.figure(figsize=(16,9))
    sns.clustermap(new_df, figsize=(20,10),cmap='coolwarm', linewidths=0.5, col_cluster=False)
    # 添加彩色条
    plt.axhspan(ymin=0, ymax=1, xmin=0, xmax=0.5, color='green', alpha=0.5)  # 组别1的绿色条
    plt.axhspan(ymin=0, ymax=1, xmin=0.5, xmax=1, color='red', alpha=0.5) 
    plt.savefig(analysis_save + xiangxi + '/' + "DEG/" + str(best_combination)[2:-2] + '_Heatmap_plot.tiff',dpi=300,bbox_inches='tight')
    plt.show()
    
    return best_result, best_combination, new_expset_combination

def Cluster_PCA_analysis(analysis_save, xiangxi, best_combination, new_expset_combination, new_expset):
    new_expset = new_expset_combination
    #凝聚层次聚类分析
    plt.figure(figsize=(15, 7))  # 设置画布尺寸
    plt.subplot(1, 2, 1)  # 创建子图，1行2列，第1列
    plt.title("Cluster Analysis")
    dend = shc.dendrogram(shc.linkage(new_expset.T, method='ward'),
                          labels=new_expset.columns)
    plt.xticks(rotation=45)  # 设置横坐标标签斜着显示
    # 显示图形
    plt.savefig(analysis_save + xiangxi + '/' + "DEG/" + str(best_combination)[2:-2] + '_Cluster_analysis.tiff',dpi=300,bbox_inches='tight')
    
    # PCA分析
    plt.subplot(1, 2, 2)  # 创建子图，1行2列，第2列
    pca = PCA(n_components=2)  # 设置维度为2
    reduced_X = pca.fit_transform(new_expset.T)
    plt.scatter([x[0] for x in reduced_X[0:3]], [x[1] for x in reduced_X[0:3]], c="r")
    plt.scatter([x[0] for x in reduced_X[3:6]], [x[1] for x in reduced_X[3:6]], c="g")
    plt.xlabel('PC1')
    plt.ylabel('PC2')
    plt.legend(['treatment', 'control'])
    plt.title('PCA')
    plt.savefig(analysis_save + xiangxi + '/' + "DEG/" + str(best_combination)[2:-2] + '_PCA_analysis.tiff',dpi=300,bbox_inches='tight')

    return

def GO_ANALYSIS_DRAW(up):
    # up 是你的数据框，包含 "term"、"n_genes" 和 "p" 列
    # 根据 "p" 列的数值大小生成颜色深度列表
    norm = plt.Normalize(up["LogP"].min(), 
                         up["LogP"].max()) 
    cmap = plt.cm.get_cmap("coolwarm")
    # 绘制气泡图
    plt.figure(figsize=(10, 6))
    sns.barplot(data=up, 
                x="n_genes", y="term", 
                hue="LogP", 
                palette=cmap,
                legend="brief")
    plt.xlabel("Number of Genes")
    plt.ylabel("GO term Name")
    plt.title("GO Enrichment of Up-regulated Genes")
    # 创建图例
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, orientation="horizontal")
    cbar.set_label("LogP")
    plt.legend(title="LogP", bbox_to_anchor=(1.05, 1), loc='upper left') 
   
    return plt

def KEGG_ANALYSIS(df_v, analysis_save, xiangxi):
    gene_sets = 'KEGG_2021_Human'
    gene_list = df_v
    enr = gp.enrichr(gene_list=gene_list,#所需查询gene_list，可以是一个列表，也可为文件（一列，每行一个基因）
                     gene_sets=gene_sets,#gene set library，多个相关的gene set 。如所有GO term组成一个gene set library.
                     organism='Human',#持(human, mouse, yeast, fly, fish, worm)， 自定义gene_set 则无影响。
                     description='kegg',#工作运行描述
                     outdir='/Users/Rui/anaconda3/envs/Bio_analysis/PY_documents/enrichr',#输出目录
                     top_term=20,
                     cutoff=0.5#pvalue阈值
                     )
    
    KEGG_RESULT = enr.results
    KEGG_RESULT["Overlap%"] = KEGG_RESULT["Overlap"].apply(lambda x: eval(x))
    KEGG_RESULT_10 = KEGG_RESULT.head(20) 
    
    norm = plt.Normalize(KEGG_RESULT_10["Adjusted P-value"].min(), 
                         KEGG_RESULT_10["Adjusted P-value"].max()) 
    cmap = plt.cm.get_cmap("coolwarm")
    # 绘制气泡图
    plt.figure(figsize=(10, 6))
    sns.scatterplot(data=KEGG_RESULT_10, 
                    x="Combined Score", y="Term", 
                    size="Overlap%", hue="Adjusted P-value", 
                    palette=cmap,
                    sizes=(50, 2000), legend="brief")
    plt.xlabel("Combined Score")
    plt.ylabel("KEGG Pathway Name")
    plt.title("KEGG Enrichment of Down-regulated Genes")
    # 创建图例
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, orientation="horizontal")
    cbar.set_label("Adjusted P-value")
    plt.legend(title="Adjusted P-value", bbox_to_anchor=(1.05, 1), loc='upper left') 
  
    return plt, KEGG_RESULT


def dict_subcellular_create(file_pathway):
    full_table = pd.read_csv(file_pathway, index_col=0, encoding='latin1', delimiter='\t')
    # 将每个键映射为一个值列表
    species_dict = {}
    # 遍历DataFrame的每一行
    for index, row in full_table.iterrows():
        # 获取行中的物种、亚细胞位置和基因符号
        species = row['Species']
        localization = row['SubCellular_Localization']
        gene_symbol = row['Gene_symbol']
        # 如果物种不在字典中，创建一个空字典
        if species not in species_dict:
            species_dict[species] = {}
        # 如果亚细胞位置不在物种字典中，创建一个空列表
        if localization not in species_dict[species]:
            species_dict[species][localization] = []
        # 将基因符号添加到相应的亚细胞位置列表中
        species_dict[species][localization].append(gene_symbol)
    
    return species_dict


def analyze_gene_localization(gene_list, species, subcellular_localizations):
    df = pd.DataFrame(columns=['Structure', 'Percentage_in_Structure', "Percentage_in_DEG", 'Appeared_Genes'])
    max_percentage = float('-inf')
    
    for structure, genes in subcellular_localizations.items():
        appeared_genes = [gene for gene in gene_list if gene in genes]
        percentage = len(appeared_genes) / len(gene_list) * 100
        percentage_in_structure = len(appeared_genes) / len(genes) * 100
        
        df = df.append({'Structure': structure,
                        'Percentage_in_Structure': percentage_in_structure,
                        'Percentage_in_DEG': percentage,
                        'Appeared_Genes': appeared_genes},
                       ignore_index=True)
        
        if structure == "Mitochondrion":
            max_percentage = percentage
            set1 = set(genes)
            set2 = set(gene_list)
    plt.figure()
    venn2([set1, set2], ('Mitochondrion', 'Gene List')) 
    plt.title("Venn Diagram") 
    # Manually adjust the positions of the numbers inside the circles 
    #plt.gca().texts[0].set_position((0.1, 0.1)) 
    #plt.gca().texts[1].set_position((0.9, 0.1)) 
    #plt.gca().texts[2].set_position((0.5, 0.9)) 
    
    return df, plt

from Bio import Entrez

def get_taxonomy_id(species_name):
    Entrez.email = "1766847007@qq.com"  # 设置你的邮箱地址
    handle = Entrez.esearch(db="taxonomy", term=species_name)
    record = Entrez.read(handle)
    if record["IdList"]:
        taxonomy_id = record["IdList"][0]
        return taxonomy_id
    else:
        return None

def get_species_name(taxonomy_id):
    Entrez.email = "your@email.com"  # 设置你的邮箱地址
    handle = Entrez.efetch(db="taxonomy", id=taxonomy_id, retmode="xml")
    record = Entrez.read(handle)
    if record:
        species_name = record[0]['ScientificName']
        return species_name
    else:
        return None


def highlight_rows_with_keywords(file_path, output_file_path, keywords):
    # Load the Excel file
    wb = load_workbook(file_path)
    ws = wb.active

    # Define the fill style for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Iterate through all rows in the worksheet
    for row in ws.iter_rows():
        # Check if any cell in the row contains any of the specified keywords
        if any(keyword.lower() in str(cell.value).lower() for cell in row for keyword in keywords):
            # Highlight all cells in the row
            for cell in row:
                cell.fill = highlight_fill

    # Save the modified Excel file
    wb.save(output_file_path)

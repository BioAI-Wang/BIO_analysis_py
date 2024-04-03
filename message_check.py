#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 29 11:47:03 2024

@author: Rui
"""
import os
import re
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
import os
import urllib.request    
import gzip
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime

def output_file_find_and_open(this_program_path, current_directory, progress):
    folder_path = this_program_path
    # 用于存储包含 "output" 的文件路径的列表
    output_files = []
    # 遍历文件夹中的所有文件和子文件夹
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            # 检查文件名是否包含 "output"
            if 'output' in file_name:
                # 构造文件的完整路径
                file_path = os.path.join(root, file_name)
                # 将路径添加到列表中
                output_files.append(file_path)
    # 打印包含 "output" 的文件路径列表
    message_output_name = output_files[0]
    # 打开 Excel 文件
    workbook = load_workbook(message_output_name)
    # 选择第一个工作表
    sheet = workbook.active
    highlighted_rows = []
    # 遍历高亮的行
    for row in sheet.iter_rows():
    # 假设高亮的标志是单元格背景色为黄色
        highlighted = any(cell.fill.start_color.index == 'FFFFFF00' for cell in row)
    # 如果该行被高亮，则读取与“GSE”列交叉处的数据
        if highlighted:
            gse_value = row[0].value  # 假设“GSE”列是第一个列
            row_data = [cell.value for cell in row]
            highlighted_rows.append((gse_value, row_data))

    # 打印结果
    for gse_value, row_data in highlighted_rows:
        print("GSE:", gse_value)
        print("Row data:", row_data)
    
    return

        

def gds_result_file_check(gds_file_path):
    # 打开文件并读取内容
    with open(gds_file_path, "r") as file:
        text = file.readlines()
    # 使用正则表达式找到Accession编号
    GSM_exsit = []
    for line in text:
        try:
            accession_match = re.search(r'Accession: GSE(\d+)', line)
        
            if accession_match:
                accession_number = accession_match.group(1)
                GSM_exsit.append("GSE"+accession_number)
        except:
            pass
    
    return GSM_exsit

def get_platform_info(geo_id):
    GSE_GPL_dict = {}
    GSE_PMID_dict ={}
    #flag = 0
    url = f'https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={geo_id}'
    # 使用Beautiful Soup解析HTML代码
    response = requests.get(url)
    # 检查响应状态码
    search_string = "Data table header descriptions"
    if response.status_code == 200:
        # 使用Beautiful Soup解析HTML代码
        soup = BeautifulSoup(response.text, 'html.parser')
        # 找到包含目标链接的<a>元素
        link_elements = soup.find_all('a')
        # 提取链接
        for link_element in link_elements:
            href = link_element.get('href')
            if href and 'GPL' in href:
                link = "https://www.ncbi.nlm.nih.gov" + href
                #print("提取的链接:", link)
                response = requests.get(link)
                # 检查响应状态码
                if response.status_code == 200:
                    if search_string in response.text:
                        gpl_id = href.split('=')[-1]
                        if geo_id not in GSE_GPL_dict: 
                            GSE_GPL_dict[geo_id] = [] 
                        # 将GPL ID添加到列表中 
                        GSE_GPL_dict[geo_id].append(gpl_id)
                        for link_element in link_elements:
                            href = link_element.get('href')
                            if href and 'pubmed' in href and "?" not in href:
                                #print(href)
                                if geo_id not in GSE_PMID_dict: 
                                    GSE_PMID_dict[geo_id] = [] 
                                # 将GPL ID添加到列表中 
                                GSE_PMID_dict[geo_id].append(href.split('/')[-1])
                                break
                            # 如果只需要第一个链接，可以添加 break 语句以提高效率
                break
    return geo_id, GSE_GPL_dict, GSE_PMID_dict

def matrix_download(database_numbers, current_directory, progress):
    try:
        folder_path = current_directory+'/'+progress+'/'+'matrix'+'/'
        # 使用 os 模块的 mkdir 函数新建文件夹
        os.mkdir(folder_path)
    except:
        pass
    for number in database_numbers:
        ftp_link = f"ftp://ftp.ncbi.nlm.nih.gov/geo/series/{number[:-3]}nnn/{number}/matrix/{number}_series_matrix.txt.gz"
        # 下载文件保存路径
        #save_path = f"downloads/{number}_series_matrix.txt.gz"  # 保存路径示例，根据需要修改
        save_path = current_directory+'/'+progress+'/'+'matrix'+'/'+f"{number}_series_matrix.txt.gz"
        # 下载文件
        try:
            urllib.request.urlretrieve(ftp_link, save_path)
            print(f"Downloaded {number} successfully.")
        except Exception as e:
            print(f"Error downloading {number}: {e}")
    print('Matrixs downloading was done!')
    
    
def matrix_check(database_numbers, current_directory, progress):
    Both_platform_matrix_exist = []
    for number in database_numbers:
        filename = current_directory+'/'+progress+'/'+'matrix'+'/'+f"{number}_series_matrix.txt.gz"
        try:
            with gzip.open(filename, 'rt') as f:
                content = f.readlines()
            if content[-3] != "!series_matrix_table_begin\n":
                Both_platform_matrix_exist.append(number)
        except:
            print(f"{number}出现未知错误")
            continue
    return Both_platform_matrix_exist

def summary_of_PMID_PID(Both_platform_matrix_exist, current_directory, progress, GSE_GPL_dict, GSE_PMID_dict):
    df1 = pd.DataFrame(Both_platform_matrix_exist, columns=["GSE"])
    df2 = pd.DataFrame(GSE_GPL_dict.items(), columns=['GSE', 'GPL'])
    df3 = pd.DataFrame(GSE_PMID_dict.items(), columns=['GSE', 'PubMed_ID'])
    merged_df = pd.merge(df1, df2, on='GSE', how='left')
    merged_df = pd.merge(merged_df, df3, on='GSE', how='left')
    merged_df.set_index('GSE', inplace=True)
    merged_df["Title"] = None
    merged_df["Journal"] = None
    for i in list(merged_df['PubMed_ID']):
        url = f'https://pubmed.ncbi.nlm.nih.gov/{i}/'
        # 使用Beautiful Soup解析HTML代码
        response = requests.get(url)
        # 检查响应状态码
        if response.status_code == 200:
            # 使用Beautiful Soup解析HTML代码
            soup = BeautifulSoup(response.text, 'html.parser')
            h1_tag = soup.find('h1', class_='heading-title')
            # 如果找到了<h1>标签，提取其中的文本内容
            if h1_tag:
                title_text = h1_tag.text.strip()
                merged_df.loc[i, 'Title'] = title_text
                
            button_element = soup.find('button', {'id': 'full-view-journal-trigger'})
            # 如果找到了按钮元素，提取其中的文本内容
            if button_element:
                button_text = button_element.text.strip()
                merged_df.loc[i, 'Journal'] = button_text
    
    merged_df = merged_df.reset_index()
    # 获取当前日期
    current_date = datetime.now().strftime('%Y_%m_%d_')
    # 生成文件名
    file_name = current_directory+'/'+progress+'/' + current_date + progress + '_output.xlsx'
    # 保存 DataFrame 到 Excel 文件
    merged_df.to_excel(file_name, index=False)
    return file_name





















